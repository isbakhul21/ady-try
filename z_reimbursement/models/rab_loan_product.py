from docutils.nodes import contact
from jsonschema.exceptions import relevance
from odoo import api, fields, models, _
from datetime import datetime, date, timedelta
from odoo.exceptions import AccessError, UserError, ValidationError

# library excel
import pandas as pd
from openpyxl import Workbook
from collections import defaultdict
import xlsxwriter
import base64
import io

def get_column_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def format_date_with_month_name(date_obj):
    result = ''
    if date_obj:
        result = date_obj.strftime('%d-%B-%Y')
    return result

def generate_months(z_calculate_month, z_start_month, z_year):
    start_date = datetime(z_year, z_start_month, 1)
    months = []
    for i in range(z_calculate_month):
        month = start_date.strftime("%B")
        year = start_date.year
        months.append(f"{month} {year}")
        start_date += timedelta(days=31)
        start_date = start_date.replace(day=1)
    return months


class RabLoanProduct(models.TransientModel):
    _name = "rab.loan.product"
    _description = "Rab Loan Product"
    _order = "id desc"

    z_partner_id = fields.Many2one('res.partner',string='Customer',tracking=True)
    z_amount_percent_rab = fields.Float(string="Persen RAB",tracking=True)

    def action_generate_loan_product(self):
        print('action_generate product Loan')
        if int(self.z_amount_percent_rab) == 0:
            raise ValidationError("Percent RAB Cannot Be Equal to 0")
        timestamp = round(datetime.now().timestamp(), 3)
        xls_filename = f'RAB-Loan-{timestamp}.xlsx'
        with io.BytesIO() as output:
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('RAB - {}'.format(self.z_partner_id.name))
            # config column
            title_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': 'yellow'})
            title_desc_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
            role_format = workbook.add_format({'border': 1,'align': 'center', 'valign': 'vcenter'})
            currency_format = workbook.add_format({'num_format': '"Rp" #,##0.00', 'border': 1})
            currency_contract_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'num_format': '"Rp" #,##0.00', 'border': 1})
            currency_total_format = workbook.add_format({'num_format': '"Rp" #,##0.00', 'border': 1, 'bg_color': 'yellow'})
            worksheet.merge_range('B2:G2', 'REPORTING RAB ADIDATA', title_format)
            worksheet.merge_range('B3:G3', 'Customer By : {}'.format(self.z_partner_id.name), title_desc_format)
            wrap_text_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
            # config length data
            length_no_internal_contract = 0
            length_no_external_contract = 0
            length_project_name = 0
            length_jangka_waktu_contract = 0
            length_nilai_contract = 0
            # header content
            worksheet.write('B5', "NO", title_format)
            worksheet.write('C5', "NO CONTRACT INTERNAL", title_format)
            worksheet.write('D5', "NO CONTRACT EXTERNAL", title_format)
            worksheet.write('E5', "PROJECT NAME", title_format)
            worksheet.write('F5', "JANGKA WAKTU KONTRAK", title_format)
            worksheet.write('G5', "NILAI KONTRAK", title_format)
            worksheet.write('H5', "DESKRIPSI", title_format)

            if length_no_internal_contract < len('NO CONTRACT INTERNAL'):
                length_no_internal_contract = len('NO CONTRACT INTERNAL')
            if length_no_external_contract < len('NO CONTRACT EXTERNAL'):
                length_no_external_contract = len('NO CONTRACT EXTERNAL')
            if length_project_name < len('PROJECT NAME'):
                length_project_name = len('PROJECT NAME')
            if length_jangka_waktu_contract < len('JANGKA WAKTU KONTRAK'):
                length_jangka_waktu_contract = len('JANGKA WAKTU KONTRAK')
            if length_nilai_contract < len('NILAI KONTRAK'):
                length_nilai_contract = len('NILAI KONTRAK')

            # checking data contract
            data_contract_ids = []
            contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_state','not in',('closed','cancel'))])
            # contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_name','in',('CT/2024/11/00112','CT/2024/11/00111','CT/2024/11/00113')),('z_state','not in',('closed','cancel'))])
            # contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_name','=','CT/2024/07/00048'),('z_state','not in',('closed','cancel'))])
            # contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_name','=','CT/2024/10/00108'),('z_state','not in',('closed','cancel'))])

            for contract_id in contract_ids:
                journal_ids = str(contract_id.z_journal_id.name).split('-')
                if len(journal_ids) > 1 and journal_ids[1].strip(' []') == 'ESCROW':
                    data_contract_ids.append(contract_id.id)
            if not data_contract_ids:
                raise ValidationError('NO data display')

            month_shelter = []
            contract_ids = contract_ids.filtered(lambda x: x.id in data_contract_ids)
            for contract_id in contract_ids.z_line_ids:
                res = generate_months(contract_id.z_calculate_month, int(contract_id.z_month), int(contract_id.z_year))
                month_shelter.extend(res)
            month_shelter_unique = list(set(month_shelter))
            month_shelter_sorted = sorted(month_shelter_unique, key=lambda x: datetime.strptime(x, '%B %Y'))

            month_start_row = 4
            month_start_column = 8
            for month in month_shelter_sorted:
                worksheet.write(month_start_row, month_start_column, month, title_format)
                month_start_column += 1


            #Start Calculate Contract
            contract_increment = 0
            contract_cell_increment = 6
            product_cell_increment = 6
            grand_total = []  # data amount akan ditambahkan di sini
            data_product_amount = []
            data_object_contract_gross = []

            for contract_id in contract_ids:
                if contract_id.z_line_ids and contract_id.z_line_ids.filtered(lambda x: x.z_sale_order_id):
                    sale_order_ids = contract_id.z_line_ids.mapped('z_sale_order_id').filtered(
                        lambda x: x.z_termin_payment_ids.filtered(lambda term: not term.z_invoice_id or term.z_invoice_id and term.z_invoice_id.payment_state != 'paid'))

                    # Menggunakan set untuk menghindari produk duplikat
                    data_product_variant = set()
                    # print("data_product_amount",data_product_variant)

                    # Kumpulkan semua product_id.id dari sale_order_id.order_line, termasuk yang duplikat
                    for sale_order_id in sale_order_ids:
                        for line_id in sale_order_id.order_line:
                            product_id = line_id.product_id.name
                            # print("product_id",product_id)
                            data_product_variant.add(product_id)  # Menambahkan ke set

                    # Tentukan jumlah baris sesuai dengan jumlah produk (tanpa duplikat)
                    length_product = len(data_product_variant)
                    # print("length_product",length_product)
                    if length_product:
                        contract_increment += 1

                        # Pengaturan jika terdapat lebih dari satu produk
                        if length_product > 1:
                            cell_row = f'B{contract_cell_increment}:B{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_increment, title_desc_format)
                            cell_row = f'C{contract_cell_increment}:C{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_name, title_desc_format)
                            cell_row = f'D{contract_cell_increment}:D{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_no_contract_external, title_desc_format)

                            cell_row = f'E{contract_cell_increment}:E{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_project_name, wrap_text_format)

                            jangka_waktu_kontrak = '{} - {}'.format(
                                format_date_with_month_name(contract_id.z_start_date),
                                format_date_with_month_name(contract_id.z_end_date)
                            )
                            cell_row = f'F{contract_cell_increment}:F{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, jangka_waktu_kontrak, title_desc_format)
                            cell_row = f'G{contract_cell_increment}:G{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_amount_total, currency_contract_format)

                            # Tambahkan sesuai jumlah produk (tanpa duplikat)
                            contract_cell_increment = contract_cell_increment + length_product - 1

                        # Pengaturan jika hanya terdapat satu produk
                        if length_product == 1:
                            cell_row = f'B{contract_cell_increment}'
                            worksheet.write(cell_row, contract_increment, title_desc_format)
                            cell_row = f'C{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_name, title_desc_format)
                            cell_row = f'D{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_no_contract_external, title_desc_format)

                            cell_row = f'E{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_project_name, wrap_text_format)

                            jangka_waktu_kontrak = '{} - {}'.format(
                                format_date_with_month_name(contract_id.z_start_date),
                                format_date_with_month_name(contract_id.z_end_date)
                            )
                            cell_row = f'F{contract_cell_increment}'
                            worksheet.write(cell_row, jangka_waktu_kontrak, title_desc_format)
                            cell_row = f'G{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_amount_total, currency_contract_format)

                        # Tambahkan baris setelah proses perhitungan untuk kontrak selesai
                        contract_cell_increment += 5
                    #End Calculate Contratct

                    #SO per Month
                    #ingat nanti data_object_contract di taro di awal loop
                    for sale_order_id in sale_order_ids:
                        total_invoice_amount = 0
                        total_paid_invoice = 0
                        for term in sale_order_id.z_termin_payment_ids:
                            total_invoice_amount += term.z_fixed_amount
                            if term.z_invoice_status == 'paid':
                                total_paid_invoice += term.z_invoice_amount_total
                        if total_invoice_amount > 0:
                            paid_percentage = (total_paid_invoice / total_invoice_amount)
                        else:
                            paid_percentage = 0

                        relevance_order_lines = sale_order_id.order_line
                        data_contract = [contract.z_contract_id.id for contract in sale_order_id]
                        data_sales_order = [sale_order.name for sale_order in sale_order_id]
                        data_product = [product.product_id.name for product in relevance_order_lines]
                        data_product_final = [data for data in data_product]
                        data_product_price = [product_price.z_price_total_contract for product_price in relevance_order_lines]
                        term_paid_percentage = paid_percentage
                        price_paid = [price * term_paid_percentage for price in data_product_price]
                        gross_price = [price - paid for price, paid in zip(data_product_price, price_paid)]
                        z_calculate_month = term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_calculate_month
                        z_month = term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_month
                        z_year = term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_year
                        product_month_year = generate_months(z_calculate_month, int(z_month), int(z_year))
                        final_price = [price / z_calculate_month if z_calculate_month > 0 else 0 for price in gross_price]

                        for product, price in zip(data_product, final_price):
                            data_object_contract_gross.append([data_contract,data_sales_order, [product, price, product_month_year]])


            print("month_shelter_sorted",month_shelter_sorted)
            # print("data_object_contract_gross",data_object_contract_gross)
            # for datas in data_object_contract_gross:
            #     print("DATAS PR",datas[2][1])
            #     print("Tyepe",type(datas[2][1]))
            # for month_year in data_object_contract_gross[2][2][2]:
            #     print("MONTH YEAR", month_year)

            # Mengelompokkan data berdasarkan kontrak, produk, dan bulan
            grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

            for entry in data_object_contract_gross:
                contract_id = entry[0][0]  # Ambil ID kontrak
                product_name = entry[2][0]  # Ambil nama produk
                product_price = entry[2][1]  # Ambil harga produk
                product_months = entry[2][2]  # Ambil bulan produk

                # Menjumlahkan harga produk berdasarkan bulan
                for month in product_months:
                    grouped_data[contract_id][product_name][month] += product_price

            # Menulis ke Excel
            index_vertical = 5
            index_horizontal = 8
            previous_contract_id = None  # Variabel untuk menyimpan contract_id sebelumnya

            grand_total_monthly = defaultdict(float)

            for contract_id, products in grouped_data.items():
                # Jika contract_id berubah, tambahkan spasi 3 sel
                if previous_contract_id is not None and previous_contract_id != contract_id:
                    index_vertical += 3  # Menambah 3 sel ke bawah

                monthly_totals = defaultdict(float)


                for product_name, month_data in products.items():
                    index_vertical += 1
                    cell_row = f'H{index_vertical}'
                    worksheet.write(cell_row, product_name, role_format)  # Menulis nama produk

                    # Menulis harga total ke Excel untuk setiap bulan
                    for month_year in month_shelter_sorted:
                        price_product_final = month_data.get(month_year, 0)
                        month_shelter_index = month_shelter_sorted.index(month_year) + 1 + index_horizontal
                        cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                        worksheet.write(cell_combine, price_product_final, currency_format)  # Menulis total harga

                        monthly_totals[month_year] += price_product_final

                        grand_total_monthly[month_year] += price_product_final

                # Menambahkan label "PLAN SUB TOTAL" di akhir setiap kontrak
                index_vertical += 1
                subtotal_cell_row = f'H{index_vertical}'
                worksheet.write(subtotal_cell_row, "SUB TOTAL", title_format)
                rab_cell_row = f'H{index_vertical + 1}'
                worksheet.write(rab_cell_row, f'RAB SUB TOTAL {self.z_amount_percent_rab} %', title_format)



                for month_year in month_shelter_sorted:
                    subtotal_value = monthly_totals.get(month_year, 0)



                    month_shelter_index = month_shelter_sorted.index(month_year) + 1 + index_horizontal
                    subtotal_cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                    worksheet.write(subtotal_cell_combine, subtotal_value, currency_total_format)
                    plan_rab_sub_total = self.z_amount_percent_rab / 100 * subtotal_value
                    rab_cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical + 1}'
                    worksheet.write(rab_cell_combine, plan_rab_sub_total, currency_total_format)



                previous_contract_id = contract_id  # Update contract_id sebelumnya

            index_vertical += 4 # Menambah 2 sel ke bawah
            grand_total_cell_row = f'H{index_vertical}'
            worksheet.write(grand_total_cell_row, "GRAND TOTAL", title_format)  # Menulis label "GRAND TOTAL"
            rab_cell_row = f'H{index_vertical + 1}'
            worksheet.write(rab_cell_row, f'RAB GRAND TOTAL {self.z_amount_percent_rab} %', title_format)

            for month_year in month_shelter_sorted:
                grand_total_value = grand_total_monthly.get(month_year, 0)
                month_shelter_index = month_shelter_sorted.index(month_year) + 1 + index_horizontal
                grand_total_cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                worksheet.write(grand_total_cell_combine, grand_total_value, currency_total_format)
                rab_grand_total = self.z_amount_percent_rab / 100 * grand_total_value
                rab_grand_total_cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical + 1}'
                worksheet.write(rab_grand_total_cell_combine, rab_grand_total, currency_total_format)



















                # print("data Contract", data_contract)
                        # print("data data_sales_order", data_sales_order)
                        # print("data Porduct", data_product)
                        # print("data data_product_price", data_product_price)
                        # print("data term_paid_percentage", term_paid_percentage)
                        # print("data price_paid", price_paid)
                        # print("data gross_price", gross_price)
                        # print("data product_month_year", product_month_year)
                        # print("data final_price", final_price)
                        # print("data data_product_final", data_product_final)

                    # print("data_object_contract",data_object_contract)
















            worksheet.set_column('C:C', length_no_internal_contract + 5)
            worksheet.set_column('D:D', length_no_external_contract + 5)
            worksheet.set_column('E:E', length_project_name + 15)
            worksheet.set_column('F:F', length_jangka_waktu_contract + 12)
            worksheet.set_column('G:G', length_nilai_contract + 8)
            worksheet.set_column('H:H', 30)
            worksheet.set_column(f'I:{get_column_letter(1000)}', 30)
            workbook.close()
            output.seek(0)
            xls_data = output.read()
        attachment_id = self.env['ir.attachment'].create({
            'name': xls_filename,
            'datas': base64.b64encode(xls_data),
            'type': 'binary',
            'res_model': self._name,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment_id.id}?download=true',
            'target': 'self',
        }

