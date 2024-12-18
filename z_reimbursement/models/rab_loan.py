from docutils.nodes import contact
from odoo import api, fields, models, _
from datetime import datetime, date, timedelta
from odoo.exceptions import AccessError, UserError, ValidationError

# library excel
import pandas as pd
from openpyxl import Workbook
import xlsxwriter
import base64
import io

def get_column_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def format_date_with_month_name(date_obj):
    result = ''
    if date_obj:
        result = date_obj.strftime('%d-%B-%Y')
    return result

def generate_months(z_calculate_month, z_start_month, z_year):
    start_date = datetime(z_year, z_start_month, 1)
    months = []
    for i in range(z_calculate_month):
        month = start_date.strftime("%B")
        year = start_date.year
        months.append(f"{month} {year}")
        start_date += timedelta(days=31)
        start_date = start_date.replace(day=1)
    return months


class RabLoanRem(models.TransientModel):
    _name = "rab.loan.rem"
    _description = "Rab Loan Rem"
    _order = "id desc"

    z_partner_id = fields.Many2one('res.partner',string='Customer',tracking=True)
    z_amount_percent_rab = fields.Float(string="Persen RAB",tracking=True)

    def action_generate(self):
        print('action_generate')
        if int(self.z_amount_percent_rab) == 0:
            raise ValidationError("Percent RAB Cannot Be Equal to 0")
        timestamp = round(datetime.now().timestamp(), 3)
        xls_filename = f'RAB-Loan-{timestamp}.xlsx'
        with io.BytesIO() as output:
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('RAB - {}'.format(self.z_partner_id.name))
            # config column
            title_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': 'yellow'})
            title_desc_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
            role_format = workbook.add_format({'border': 1})
            currency_format = workbook.add_format({'num_format': '"Rp" #,##0.00', 'border': 1})
            currency_contract_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'num_format': '"Rp" #,##0.00', 'border': 1})
            currency_total_format = workbook.add_format({'num_format': '"Rp" #,##0.00', 'border': 1, 'bg_color': 'yellow'})
            worksheet.merge_range('B2:G2', 'REPORTING RAB ADIDATA', title_format)
            worksheet.merge_range('B3:G3', 'Customer By : {}'.format(self.z_partner_id.name), title_desc_format)
            wrap_text_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
            # config length data
            length_no_internal_contract = 0
            length_no_external_contract = 0
            length_project_name = 0
            length_jangka_waktu_contract = 0
            length_nilai_contract = 0
            # header content
            worksheet.write('B5', "NO", title_format)
            worksheet.write('C5', "NO CONTRACT INTERNAL", title_format)
            worksheet.write('D5', "NO CONTRACT EXTERNAL", title_format)
            worksheet.write('E5', "PROJECT NAME", title_format)
            worksheet.write('F5', "JANGKA WAKTU KONTRAK", title_format)
            worksheet.write('G5', "NILAI KONTRAK", title_format)
            worksheet.write('H5', "DESKRIPSI", title_format)

            if length_no_internal_contract < len('NO CONTRACT INTERNAL'):
                length_no_internal_contract = len('NO CONTRACT INTERNAL')
            if length_no_external_contract < len('NO CONTRACT EXTERNAL'):
                length_no_external_contract = len('NO CONTRACT EXTERNAL')
            if length_project_name < len('PROJECT NAME'):
                length_project_name = len('PROJECT NAME')
            if length_jangka_waktu_contract < len('JANGKA WAKTU KONTRAK'):
                length_jangka_waktu_contract = len('JANGKA WAKTU KONTRAK')
            if length_nilai_contract < len('NILAI KONTRAK'):
                length_nilai_contract = len('NILAI KONTRAK')

            # checking data contract
            data_contract_ids = []
            contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_state','not in',('closed','cancel'))])
            # contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_name','in',('CT/2024/08/00079','CT/2024/08/00078','CT/2024/08/00080')),('z_state','not in',('closed','cancel'))])
            # contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_name','=','CT/2024/08/00079'),('z_state','not in',('closed','cancel'))])
            # contract_ids = self.env['res.contract'].sudo().search([('z_partner_id','=',self.z_partner_id.id),('z_journal_id','!=',False),('z_name','=','CT/2024/10/00108'),('z_state','not in',('closed','cancel'))])
            for contract_id in contract_ids:
                journal_ids = str(contract_id.z_journal_id.name).split('-')
                if len(journal_ids) > 1 and journal_ids[1].strip(' []') == 'ESCROW':
                    data_contract_ids.append(contract_id.id)
            if not data_contract_ids:
                raise ValidationError('NO data display')

            month_shelter = []
            contract_ids = contract_ids.filtered(lambda x: x.id in data_contract_ids)
            for contract_id in contract_ids.z_line_ids:
                res = generate_months(contract_id.z_calculate_month, int(contract_id.z_month), int(contract_id.z_year))
                month_shelter.extend(res)
            month_shelter_unique = list(set(month_shelter))
            month_shelter_sorted = sorted(month_shelter_unique, key=lambda x: datetime.strptime(x, '%B %Y'))

            month_start_row = 4
            month_start_column = 8
            for month in month_shelter_sorted:
                worksheet.write(month_start_row, month_start_column, month, title_format)
                month_start_column += 1


            #Start Calculate Contract
            contract_increment = 0
            contract_cell_increment = 6
            product_cell_increment = 6
            grand_total = []  # data amount akan ditambahkan di sini
            data_product_amount = []

            for contract_id in contract_ids:
                if contract_id.z_line_ids and contract_id.z_line_ids.filtered(lambda x: x.z_sale_order_id):
                    sale_order_ids = contract_id.z_line_ids.mapped('z_sale_order_id').filtered(lambda x: x.z_termin_payment_ids.filtered(lambda term: not term.z_invoice_id or term.z_invoice_id and term.z_invoice_id.payment_state != 'paid'))
                    data_product_variant = []
                    # Kumpulkan semua product_id.id dari sale_order_id.order_line, termasuk yang duplikat
                    for sale_order_id in sale_order_ids:
                        for line_id in sale_order_id.order_line:
                            data_product_variant.append(line_id.product_id.id)
                    # Tentukan jumlah baris sesuai dengan jumlah produk (termasuk duplikat)
                    length_product = len(data_product_variant)
                    if length_product:
                        contract_increment += 1

                        # Pengaturan jika terdapat lebih dari satu produk
                        if length_product > 1:
                            cell_row = f'B{contract_cell_increment}:B{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_increment, title_desc_format)
                            cell_row = f'C{contract_cell_increment}:C{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_name, title_desc_format)
                            cell_row = f'D{contract_cell_increment}:D{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_no_contract_external, title_desc_format)

                            cell_row = f'E{contract_cell_increment}:E{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_project_name, wrap_text_format)

                            jangka_waktu_kontrak = '{} - {}'.format(
                                format_date_with_month_name(contract_id.z_start_date),
                                format_date_with_month_name(contract_id.z_end_date)
                            )
                            cell_row = f'F{contract_cell_increment}:F{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, jangka_waktu_kontrak, title_desc_format)
                            cell_row = f'G{contract_cell_increment}:G{contract_cell_increment + length_product - 1}'
                            worksheet.merge_range(cell_row, contract_id.z_amount_total, currency_contract_format)

                            # Tambahkan sesuai jumlah produk (termasuk duplikat)
                            contract_cell_increment = contract_cell_increment + length_product - 1

                        # Pengaturan jika hanya terdapat satu produk
                        if length_product == 1:
                            cell_row = f'B{contract_cell_increment}'
                            worksheet.write(cell_row, contract_increment, title_desc_format)
                            cell_row = f'C{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_name, title_desc_format)
                            cell_row = f'D{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_no_contract_external, title_desc_format)

                            cell_row = f'E{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_project_name, wrap_text_format)

                            jangka_waktu_kontrak = '{} - {}'.format(
                                format_date_with_month_name(contract_id.z_start_date),
                                format_date_with_month_name(contract_id.z_end_date)
                            )
                            cell_row = f'F{contract_cell_increment}'
                            worksheet.write(cell_row, jangka_waktu_kontrak, title_desc_format)
                            cell_row = f'G{contract_cell_increment}'
                            worksheet.write(cell_row, contract_id.z_amount_total, currency_contract_format)

                        # Tambahkan baris setelah proses perhitungan untuk kontrak selesai
                        contract_cell_increment += 4
                    #End Calculate Contratct

                    #SO Per Month Calculate
                    for sale_order_id in sale_order_ids:
                        total_invoice_amount = 0
                        total_paid_amount = 0
                        for term in sale_order_id.z_termin_payment_ids:
                            total_invoice_amount += term.z_fixed_amount
                            if term.z_invoice_status == 'paid':
                                total_paid_amount += term.z_invoice_amount_total

                        if total_invoice_amount > 0:
                            paid_percentage = (total_paid_amount / total_invoice_amount) * 100
                        else:
                            paid_percentage = 0

                        relevant_order_lines = term.z_sale_order_id.order_line
                        last_price = [price - (price * paid_percentage / 100) for price in term.z_sale_order_id.order_line.mapped('z_price_total_contract')]



                        z_calculate_month = term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_calculate_month
                        z_month = term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_month
                        z_year = term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_year
                        result_month_year = generate_months(z_calculate_month, int(z_month), int(z_year))

                        if z_calculate_month > 0:
                            price_divide_month = [price / z_calculate_month for price in last_price]
                        else:
                            price_divide_month = [0 for price in last_price]

                        amount_sales_order = sum(price_divide_month)
                        data_product = [x.product_id.name for x in relevant_order_lines]
                        data_product_price = [x.z_price_total_contract for x in relevant_order_lines]

                        data_product_amount.append(
                            (
                                paid_percentage,
                                term.z_sale_order_id.z_contract_id.id,
                                term.z_sale_order_id.name,
                                data_product,
                                data_product_price,
                                last_price,
                                term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_calculate_month,
                                term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_month,
                                term.z_sale_order_id.z_contract_id.z_line_ids.filtered(lambda t: t.z_sale_order_id.id == term.z_sale_order_id.id).z_year,
                                result_month_year,
                                amount_sales_order,
                                price_divide_month,

                            )
                        )

            # print('month_shelter_sorted',month_shelter_sorted)
            # print('data_product_amount : ', data_product_amount)


            index_horizontal = 8
            index_vertical = 5
            data_contract = []
            data_total_amount_sales = []


            for products in data_product_amount:
                result_month_year = products[9]
                contract_id = products[1]
                sales_amount = products[10]
                data_total_amount_sales.append((contract_id,result_month_year,sales_amount))
                if int(products[1]) not in data_contract:
                    if data_contract:
                        index_vertical += 3
                    data_contract.append(int(products[1]))


                for product, last_price in zip(products[3],products[11]):
                    index_vertical += 1
                    cell_row = f'H{index_vertical}'
                    worksheet.write(cell_row, product, role_format)

                    months_not_in_month_shelter_sorted = [month for month in month_shelter_sorted if month not in result_month_year]
                    for x_month in result_month_year:
                        if x_month in month_shelter_sorted:
                            month_shelter_index = month_shelter_sorted.index(x_month) + 1 + index_horizontal
                            cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                            worksheet.write(cell_combine, last_price, currency_format)

                    for x_month in months_not_in_month_shelter_sorted:
                        month_shelter_index = month_shelter_sorted.index(x_month) + 1 + index_horizontal
                        cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                        worksheet.write(cell_combine, 0, currency_format)




            index_vertical = 5
            index_horizontal = 8
            data_contract_total = []
            grand_total_per_month = {month: 0 for month in month_shelter_sorted}
            for products in data_product_amount:
                if int(products[1]) not in data_contract_total:
                    data_product = [d[3] for d in data_product_amount if products[1] == d[1]]

                    # print("data_product",data_product)
                    data_product_final = []
                    for d in data_product:
                        print("product",d)
                        for e in d:
                            data_product_final.append(e)
                    index_vertical = index_vertical + len(data_product_final) + 1
                    if data_contract_total:
                        index_vertical += 2
                    data_contract_total.append(int(products[1]))

                for x_month in products[9]:
                    if x_month in month_shelter_sorted:
                        product_amount_final = 0
                        for p in data_product_amount:
                            if p[1] == products[1] and x_month in p[9]:
                                product_amount_final += float(p[10])

                        grand_total_per_month[x_month] += product_amount_final


                        month_shelter_index = month_shelter_sorted.index(x_month) + 1 + index_horizontal
                        print('month_shelter_index : ', month_shelter_index)
                        print('index_vertical : ', index_vertical)
                        cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                        worksheet.write(cell_combine, product_amount_final, currency_total_format)

                        plan_rab_total = product_amount_final * self.z_amount_percent_rab / 100
                        cell_plan = f'{get_column_letter(month_shelter_index)}{index_vertical + 1}'
                        worksheet.write(cell_plan, plan_rab_total, currency_total_format)

                        #Label
                        cell_desc_sub_total = f'H{index_vertical}'
                        worksheet.write(cell_desc_sub_total, "SUB TOTAL", currency_total_format)
                        cell_desc_sub_total = f'H{index_vertical + 1}'
                        worksheet.write(cell_desc_sub_total, "PLAN RAB SUB TOTAL", currency_total_format)

                        # months_not_in_x_month = [month for month in month_shelter_sorted if month not in products[9]]
                        # for missing_month in months_not_in_x_month:
                        #     month_shelter_index = month_shelter_sorted.index(missing_month) + 1 + index_horizontal
                        #     cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                        #     worksheet.write(cell_combine, 0, currency_total_format)
                        #     cell_plan_sub_total = f'{get_column_letter(month_shelter_index)}{index_vertical + 1}'
                        #     worksheet.write(cell_plan_sub_total, 0, currency_total_format)







            # looping for  grand total
            index_vertical += 3
            cell_desc_grand_total = f'H{index_vertical}'
            worksheet.write(cell_desc_grand_total, "GRAND TOTAL", currency_total_format)
            cell_desc_grand_total = f'H{index_vertical + 1}'
            worksheet.write(cell_desc_grand_total, "PLAN RAB GRAND TOTAL", currency_total_format)


            for month, total in grand_total_per_month.items():
                month_shelter_index = month_shelter_sorted.index(month) + 1 + index_horizontal
                cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical}'
                worksheet.write(cell_combine, total, currency_total_format)

                cell_combine = f'{get_column_letter(month_shelter_index)}{index_vertical + 1}'
                grand_total = total * self.z_amount_percent_rab / 100
                worksheet.write(cell_combine, grand_total, currency_total_format)









            worksheet.set_column('C:C', length_no_internal_contract + 5)
            worksheet.set_column('D:D', length_no_external_contract + 5)
            worksheet.set_column('E:E', length_project_name + 15)
            worksheet.set_column('F:F', length_jangka_waktu_contract + 12)
            worksheet.set_column('G:G', length_nilai_contract + 8)
            worksheet.set_column('H:H', 30)
            worksheet.set_column(f'I:{get_column_letter(1000)}', 30)
            workbook.close()
            output.seek(0)
            xls_data = output.read()
        attachment_id = self.env['ir.attachment'].create({
            'name': xls_filename,
            'datas': base64.b64encode(xls_data),
            'type': 'binary',
            'res_model': self._name,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment_id.id}?download=true',
            'target': 'self',
        }

